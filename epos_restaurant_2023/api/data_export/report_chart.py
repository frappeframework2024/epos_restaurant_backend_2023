
import frappe
from frappe.desk import query_report
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference,BarChart,PieChart,DoughnutChart
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from frappe import _
import string
import random
from openpyxl.chart.label import DataLabelList
from io import BytesIO
from epos_restaurant_2023.api.data_export.utils import cell_array


def render_chart_data(ws2,data):

    # set cell a1
    columns = ["Series"] + [d.get("name") for d in data.get("data").get("datasets")]
   
    for index,c in enumerate(columns):
        ws2.cell(row=1, column=index + 1, value=c)

    # loop label and get data
    for index,l in enumerate(data.get("data").get("labels")):
        if data.get("type") in ["pie","donut"]:
            
            ws2.cell(row=index + 2, column=1,value= "{} ({})".format(l,data.get("data").get("datasets")[0]["values"][index]))
        else:
            ws2.cell(row=index + 2, column=1,value= l )
        for j,d in enumerate(data.get("data").get("datasets")):
            ws2.cell(row=index + 2, column=j+2,value= d.get("values")[index])
            

        

def render_chart(ws2,chart_data,columns):

    max_row = len( chart_data.get("data").get("labels"))
    datasets = chart_data.get("data").get("datasets")

    if chart_data.get("type")=="line":
        chart = LineChart()
    elif chart_data.get("type") =="bar": 
        chart = BarChart()
    elif chart_data.get("type") =="pie": 
        chart =  PieChart()    
    else:#donut 
        chart =  DoughnutChart()

    
    
    chart.width =len(columns)*2#I want to set with to extimant of total report column 
    chart.height = 9

    for index,d in enumerate( datasets):
        data_reference = Reference(ws2, min_col=2+index, min_row=1, max_row=max_row +1)  
        chart.add_data(data_reference, titles_from_data=True)

    categories_reference = Reference(ws2, min_col=1, min_row=2, max_row=max_row+1)  # chart series
    chart.set_categories(categories_reference)

    # set label and value over point
    # each chart type difference setting
    if chart_data.get("type")=="line":
        for series in chart.series:
            series.graphicalProperties.line.width = 12000    # Approx. 2px line width
            series.smooth = True
            # Add markers (points) to the line
            series.marker.symbol = "circle"  # Use 'circle' for the marker
            series.marker.size = 7  # Set marker size (7pt)

            # Set the marker color
            series.marker.graphicalProperties.solidFill = "FFFFFF"  # White fill for the marker

            series.dLbls = DataLabelList()
            series.dLbls.showVal = True
            series.dLbls.position = 't' 

    elif chart_data.get("type")=="bar":
        chart.dataLabels = DataLabelList() 
        chart.dataLabels.showVal = True
    elif chart_data.get("type") in ["pie"]:
        chart.dataLabels = DataLabelList() 
        chart.dataLabels.showCatName = True
        chart.dataLabels.showPercent = True
        chart.dataLabels.dLblPos = 'outEnd' # Position the labels outside the pie chart
        chart.dataLabels.showLeaderLines = True  # Add leader lines from the chart to the labels
    elif chart_data.get("type") in ["donut"]:
        chart.dataLabels = DataLabelList() 
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = True
        chart.holeSize = 50
        

    
    chart_start_cell = cell_array()[len( chart_data.get("data").get("datasets")) + 6] + "2"

    ws2.add_chart(chart, chart_start_cell )
    
