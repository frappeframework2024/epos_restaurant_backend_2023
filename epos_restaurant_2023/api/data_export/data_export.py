import frappe
from frappe.desk import query_report
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference,BarChart
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Font, Alignment,PatternFill, Border, Side
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from frappe import _
import string
import random
from openpyxl.chart.label import DataLabelList
from io import BytesIO

import os
from frappe.utils import get_site_path
from epos_restaurant_2023.api.data_export import report_chart
from epos_restaurant_2023.api.data_export.utils import cell_array,get_business_information,get_status_color_by_alias
import base64
from openpyxl.drawing.image import Image
from PIL import Image as PILImage  # Import Pillow


min_chart_image_width = 1000
max_chart_image_width = 1500
max_chart_image_height = 350


setting_doc = None
report_data_row = 25
has_summary_kpi = False

@frappe.whitelist()
def export_excel(report_name="Reservation List Report",report_data=None,chart_image=None,filters_html=None,filters = None):
    global setting_doc
    global report_data_row
         
    setting_doc = frappe.get_cached_doc("Export Excel Setting", None)
    setting_doc.report_data_start_row_with_chart = 8

    if report_data.get("report_summary") :
        setting_doc.report_data_start_row_with_chart = setting_doc.report_data_start_row_with_chart +3
        setting_doc.chart_start_cell = setting_doc.chart_start_cell +3
        
    if chart_image:
        setting_doc.report_data_start_row_with_chart = setting_doc.report_data_start_row_with_chart + 18

    
    wb = Workbook()
    # Create a PageMargins object
    margins = PageMargins(top=0.3, bottom=0.3, left=0.3, right=0.3, header=0.3, footer=0.3)
    ws1 = wb.active
    
    ws1.title = "Report"
    ws1.page_margins = margins
    
    render_report_letter_head(ws1,filters,report_data["columns"])

    render_header(ws1, report_name, report_data["columns"])
    
    if filters_html:
        render_filter(ws1,filters_html, report_data["columns"])
    
    

 
    if report_data.get("chart"):
        report_data_row= setting_doc.report_data_start_row_with_chart 
        # render chart
        ws2 = wb.create_sheet(title="Chart Data")
        report_chart.render_chart_data(ws2, report_data.get("chart"))
    
        report_chart.render_chart(ws2, report_data.get("chart"),report_data["columns"])
    else:
        report_data_row= setting_doc.report_data_start_row_with_without_chart 
    
    
    #  set row column data widh
    # we need to get with to calculate summry kpi height and chart height
    for index,c in enumerate(report_data["columns"]):
        ws1.column_dimensions[cell_array()[index]].width = get_column_width(c)

 
 
    cell_width = sum([d.get("cell_width") for d in report_data["columns"]])*7
    
    if cell_width<setting_doc.chart_image_min_width:
        cell_width = setting_doc.chart_image_min_width
    if cell_width>max_chart_image_width:
        cell_width = max_chart_image_width

    if report_data.get("report_summary"):
        render_report_summary_kpi(ws1, report_summary= report_data.get("report_summary"))
 
    # render chart using report chart image
    
    if chart_image:
        if chart_image.startswith("data:image/"):
            # Find the comma and slice the string
            chart_image = chart_image.split(",")[1]

        image_data = base64.b64decode(chart_image)
        image_stream = BytesIO(image_data)
        with PILImage.open(image_stream) as img:
            original_width, original_height = img.size
        
        
        # Calculate the new height to maintain aspect ratio
        new_height = int((original_height / original_width) * cell_width)
        new_height = min(new_height, max_chart_image_height)
        
        img = Image(image_stream)
        img.width = cell_width
        img.height = new_height
        
        ws1.add_image(img, f'A{setting_doc.chart_start_cell}')






    # render report data
    render_report_data(ws1,columns=report_data["columns"], data=report_data["result"],report_data_row=report_data_row)
    
    
    if report_data["result"][0].get("report_summary"):
        start_row_index = report_data_row + len(report_data["result"]) + 2#add 2 because include 1 for row header and 1 for space 1 row
        if  report_data.get("report_summary")  and  not  report_data.get("chart"): 
            start_row_index = start_row_index + 6
            
        render_report_summary(ws1=ws1, 
                                summary_data= report_data["result"][0].get("report_summary"),
                                summary_fields= report_data["result"][0].get("report_summary_fields"),
                                start_row_index = start_row_index
                                )
    
    # save and return url

    str_name = ''.join(random.choices(string.ascii_letters,k=7)) 
    file_name = "line_chart_with_title_currency_format_{}.xlsx".format(str_name)
    file_path = os.path.join(get_site_path('public', 'files') ,"excel", file_name)
    


    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)

    # Create a file name with today's date
    filename = f"Exported_Data.xlsx"

    # Create a response to download the Excel file
    frappe.local.response.filecontent = xlsx_file.getvalue()
    frappe.local.response.type = "binary"
    frappe.local.response.filename = filename




    # wb.save(file_path)

    # # Create a URL for the file to be accessed from the browser
    # file_url = '/files/excel/' + file_name

    # # Return the URL for the file to be opened or downloaded in the browser
    # return {"file_url": file_url}

def render_report_letter_head(ws1,filters,columns):
    info = get_business_information(filters)

    row_index = 1
    total_column  =  len(columns) + 1# include No Column
    if total_column<setting_doc.min_column:
        total_column = setting_doc.min_column
    
    if total_column>setting_doc.max_column:
        total_column = setting_doc.max_column
        
        
    end_merge_cell = cell_array()[total_column - 1]  

    ws1.merge_cells(f'A{row_index}:{end_merge_cell}{row_index}') 
    header_cell = f'A{row_index}'
    ws1[header_cell] = info.get("header")
    ws1[header_cell].font = Font(size=24)  
    ws1[header_cell].alignment = Alignment(vertical='center',horizontal= "center" )


    ws1.merge_cells(f'B2:{end_merge_cell}2') 
    ws1[f"B2"] = info.get("sub_header")
    ws1[f"B2"].font = Font(size=12)  
    ws1[f'B2'].alignment = Alignment(vertical='center',horizontal="center" ,wrap_text=True)
    ws1.row_dimensions[2].height = len(info.get("sub_header").split("\n")) * 15

    # show logo if have
    if info.get("logo"):
        image_path = frappe.get_site_path('public')
        
        img = Image(image_path + info.get("logo"))
        original_height = img.height
        original_width = img.width
        
        
        # Calculate the new height to maintain aspect ratio
        aspect_ratio = original_height / original_width

        # Calculate the new height to maintain the aspect ratio
        new_height = int(setting_doc.logo_width * aspect_ratio)

        img.width = setting_doc.logo_width
        img.height =new_height
        
        ws1.add_image(img, 'A1')  # Puts the image in cell A1


def render_header(ws1, report_name,columns):
    
    row_index = setting_doc.report_title_start_row
    total_column  = len(columns) + 1# include No Column
    if total_column<setting_doc.min_column:
        total_column = setting_doc.min_column
    if total_column>setting_doc.max_column:
        total_column = setting_doc.max_column
        
    end_merge_cell = cell_array()[total_column - 1]  
    ws1.merge_cells(f'A{row_index}:{end_merge_cell}{row_index}') 
    header_cell = f'A{row_index}'
    ws1[header_cell] =  _(report_name)
    ws1[header_cell].font = Font(size=18)  
    ws1[header_cell].alignment = Alignment(vertical='center',horizontal= "center" )

def render_filter(ws1, filters_html,columns ):
    row_index = setting_doc.report_title_start_row + 1
    total_column  =  len(columns) + 1# include No Column
    if total_column<setting_doc.min_column:
        total_column = setting_doc.min_column
    if total_column>setting_doc.max_column:
        total_column = setting_doc.max_column
    ws1.row_dimensions[row_index].height =30
    ws1.merge_cells(f'A{row_index}:{cell_array()[total_column-4]}{row_index}') 
    ws1[f'A{row_index}'] = frappe.utils.strip_html(filters_html).replace("&amp;", "&")
    ws1[f'A{row_index}'].font = Font(size=10,italic=True)  
    ws1[f'A{row_index}'].alignment = Alignment(vertical='center' ,wrap_text=True)

    start_cell = cell_array()[total_column-3]
    end_cell = cell_array()[total_column-1]
    
    ws1.merge_cells(f'{start_cell}{row_index}:{end_cell}{row_index}') 
    
    printed_date =  "Printed Date: {}".format(frappe.format(frappe.utils.now(),{"fieldtype":"Datetime"}))
    ws1[f'{start_cell}{row_index}'] = _("Printed By: ") + frappe.get_cached_value("User", frappe.session.user,"full_name") + "\n" + printed_date
    
    ws1[f'{start_cell}{row_index}'].font =  Font(size=10,italic=True) 
    ws1[f'{start_cell}{row_index}'].alignment = Alignment(vertical='center',horizontal= "right" ,wrap_text=True)
    for c in range(0,total_column):
        ws1[f'{cell_array()[c]}{row_index}'].border =  Border(bottom=Side(style='thin')) 
    
    
def render_report_summary_kpi(ws1,  report_summary):
    row_index = 7
    ws1.row_dimensions[row_index+1].height =25
    ws1.row_dimensions[row_index+1].height =30
    # color = [{"red":"FF0000"},{"blue":"0000FF"},{"orange":"FFA500"},{"green":"00FF00"}]
    if report_summary:
        index = 1
        for   d in report_summary:
            label_cell = ws1.cell(row= row_index, column=index + 1, value=d.get("label"))
            value_cell = ws1.cell(row= row_index + 1, column=index + 1, value=d.get("value"))
            label_cell.alignment = Alignment(vertical='center',horizontal= "center",wrap_text=True)
            value_cell.alignment = Alignment(vertical='center',horizontal= "center" )
            value_color = ""
            if d.get("indicator","") =='blue':
                value_color = "0000FF"
            elif d.get("indicator","") =='red':
                 value_color = "FF0000"
                
            elif d.get("indicator","") =='green':
                 value_color = "228B22"
            elif d.get("indicator","") =='orange':
                 value_color = "FFA500"
            else:
                value_color = "000000"
                
                
            value_cell.font = Font(size=13, color=value_color)  
            
            index = index + 1


def render_report_data(ws1,columns,data,report_data_row=25):
    report_data_row= setting_doc.report_data_start_row_with_chart
    index =1
    # No
    ws1.column_dimensions["A"].width =5
    cell = ws1.cell(row=report_data_row, column=1, value="No")
    format_header_cell(cell,"center")
    
    for  c in columns:
        width =  get_column_width(c)
        ws1.column_dimensions[cell_array()[index]].width =width
        ws1.row_dimensions[report_data_row].height =25
        cell = ws1.cell(row=report_data_row, column=index+1, value=c.get("label"))
    # mearge cell
        if c.get("merge_cell",1)>1:
            merge_cell = f'{cell_array()[index]}{report_data_row}:{cell_array()[index   +  c.get("merge_cell")  -1 ]}{report_data_row}'
          
            ws1.merge_cells(merge_cell)
            for column  in range(index, index +  1  +  c.get("merge_cell",1)):
                ws1.column_dimensions[cell_array()[column]].width =width / c.get("merge_cell",1)
                ws1.cell(row=report_data_row, column=column+1).border =  Border(bottom=Side(style='thin')) 
            
       
                    
        format_header_cell(cell,c.get("align","left"))
        
        index = index + c.get("merge_cell",1)

     
    for row_index,d in enumerate(data):
        # No numbe order row
        if d.get("is_total_row",0) == 0:
            cell = ws1.cell(row=row_index + report_data_row + 1, column=1, value=row_index + 1 )
            cell.alignment = Alignment(vertical='center',horizontal= "center" )
            
        column_index = 1
        for c in columns:
            if str(type(d)) != "<class 'list'>":
               
                value =  d.get(c.get("fieldname"))
                value = frappe.format(value, {"fieldtype":c.get("fieldtype","Data")})
                if column_index==0:
                    # set indent space character 
                    value = f'{" " * d.get("indent",0) * 5} {value}'
                    
                cell_row = row_index + report_data_row + 1
                
                cell = ws1.cell(row= cell_row, column=column_index+1, value=value)
                    
                
                
                # mearge cell
                if c.get("merge_cell",1)>1:
                    merge_cell = f'{cell_array()[column_index]}{cell_row}:{cell_array()[column_index   +  c.get("merge_cell")  -1 ]}{cell_row}'
                    
                    ws1.merge_cells(merge_cell)
            
                cell.alignment = Alignment(vertical='center',horizontal= c.get("align","left") )
                
                # check if report is Monthly Room Availilability Chart 
                # then check hight row by reservation status
                if c.get("set_reservation_status_color"):
                    color = get_status_color_by_alias(value)
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    if color:
                        cell.fill = fill

                if d.get("is_total_row",0) ==1 or d.get("is_group",0) == 1:
                    ws1.row_dimensions[row_index + report_data_row].height =25
                    cell.font = Font(bold=True) 
                    cell.border = Border(top=Side(style='thin')) 
                    
                column_index = column_index + c.get("merge_cell",1) 
                # check if cell is group then check if it  set to merge
                # then break stop loop
                if d.get("merge_group_row",0)==1:
                        ws1.merge_cells('A{row}:{end_column}{row}'.format(row=row_index + report_data_row+1, end_column = cell_array()[len(columns)-1]))
                        break
            
            else:
                for col_num, value in enumerate(d, start=1):
                    cell = ws1.cell(row= row_index + report_data_row + 1, column=col_num, value=value) 
                    cell.font = Font(bold=True) 
           
            
def format_header_cell(cell,align):
    cell.alignment = Alignment(vertical='center',horizontal= align) 
    cell.font = Font(bold=True) 
    silver_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    cell.fill = silver_fill
    cell.border = Border(bottom=Side(style='thin'))                        

def render_report_summary(ws1, summary_data,summary_fields,start_row_index):
    
    c_index = 1
    ws1.row_dimensions[start_row_index].height =20
    for c in summary_fields :
        if c.get("merge_cell",1)>1:
            merge_cell = '{start_cell}{row_index}:{end_cell}{row_index}'.format(start_cell =  cell_array()[c_index],end_cell = cell_array()[c_index + c.get("merge_cell") -1],row_index = start_row_index )
            
            ws1.merge_cells(merge_cell)
        

        cell = ws1.cell(row=start_row_index, column=c_index+1, value=c.get("label"))
        cell.alignment = Alignment(vertical='center',horizontal= c.get("align","left") )    
        cell.font = Font(bold=True) 
        silver_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        cell.fill = silver_fill
      
        
        c_index = c_index + c.get("merge_cell",1)

    # set border of summary table header bootm
    for i in range(2, sum([d.get("merge_cell",1) for d in summary_fields]) + 2):
        cell = ws1.cell(row=start_row_index, column=i)
        cell.border = Border(bottom=Side(style='thin'))

    start_row_index = start_row_index + 1
    
    for index,d in enumerate(summary_data):
        c_index = 1
        for c in summary_fields :
            if c.get("merge_cell",1)>1:
                merge_cell = '{start_cell}{row_index}:{end_cell}{row_index}'.format(start_cell =  cell_array()[c_index],end_cell = cell_array()[c_index + c.get("merge_cell") -1],row_index = start_row_index+index )
                ws1.merge_cells(merge_cell)

            cell = ws1.cell(row=start_row_index + index, column=c_index+1, value= frappe.format(d.get(c.get("fieldname")),{"fieldtype":c.get("fieldtype","Data")}))
            cell.alignment = Alignment(vertical='center',horizontal= c.get("align","left") )    
            c_index = c_index + c.get("merge_cell",1)
            
        if d.get("is_total_row",0)==1:
            ws1.row_dimensions[start_row_index + index].height =20
            for i in range(2, sum([d.get("merge_cell",1) for d in summary_fields]) + 2):
                cell = ws1.cell(row=start_row_index + index, column=i)
                cell.font = Font(bold=True) 
                cell.border = Border(top=Side(style='thin'))




def get_column_width(column):
    
    widths = [
        {"label": "Res #","width": 12},
        {"label": "Stay #","width": 12},
        {"fieldtype": "Date","width": 11},
        {"label": "Type","width": 8}, #for reservation type, label can be type or Res. Type
        {"label": "Res. Type","width": 8}, #for reservation type, label can be type or Res. Type
        {"label": "Room","width": 9},
        {"label": "Rooms","width": 9},
        {"label": "Room Type","width": 10},
        {"label": "Room Types","width": 10},
        {"label": "Night","width": 7},
        {"label": "Nights","width": 7},
        {"label": "Pax","width": 7},
        {"label": "Pax(A/C)","width": 8},
        {"label": "Source","width": 15},
        {"label": "Business Source","width": 15},
        {"label": "Guest","width": 15},
        {"label": "Guest Name","width": 15},
        {"label": "Status","width": 10},
        {"label": "Res. Status","width": 10},
        {"fieldtype": "Currency","width": 10},
        {"label": "Room Charge","width": 12},
        {"label": "Other Charge","width": 12},
        {"label": "Total Amount","width": 12},
        {"label": "Total Charge","width": 12}, 
        {"label": "ADR Amount","width": 12}, 

    ]
    
    if column.get("cell_width",0)>0:
        return column.get("cell_width")
    
    
        
    w =  [d["width"] for d in widths if d.get("label","") == column.get("label","") ]
    # check with fieldtype
    if not w:
        w =  [d["width"] for d in widths if d.get("fieldtype","") == column.get("fieldtype","") ]

    if w:
        column["cell_width"] = w[0]
        return w[0]
    

    column["cell_width"] = column.get("width",125)/7

    return column.get("width",100)/7
